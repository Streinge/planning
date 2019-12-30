import os
from openpyxl import load_workbook


file_name = '/home/streinge/planning/planning/reports/План поток  29.11.xlsx'
# Import `load_workbook` module from `openpyxl`

from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook(file_name)

# Get sheet names
print(wb.get_sheet_names())
